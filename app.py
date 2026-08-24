import base64
import io
import json
import re
import shutil
import subprocess
import tempfile
import traceback
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter

OLLAMA_GENERATE_URL = "http://localhost:11434/api/generate"
OLLAMA_TAGS_URL = "http://localhost:11434/api/tags"
DEFAULT_MODEL = "qwen2.5vl:3b"

# -----------------------------------------------------------------------------
# Image preparation
# -----------------------------------------------------------------------------
def preprocess_image(pil_img: Image.Image, block_size: int = 25, c_value: int = 15) -> np.ndarray:
    img = np.array(pil_img.convert("RGB"))
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    thresh = cv2.adaptiveThreshold(
        denoised,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        blockSize=block_size,
        C=c_value,
    )
    kernel = np.ones((2, 2), np.uint8)
    return cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)


def prepare_vision_image_bytes(image_bytes: bytes, max_width: int = 1280) -> bytes:
    """Prepare a reasonably sized image for a local vision model.

    The previous version upscaled small scans to 2600 px and sent PNGs, which can
    dramatically increase vision-token processing time. This version never
    enlarges an image. It only downsizes very large pages and sends compact JPEG.
    """
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    if img.width > max_width:
        scale = max_width / img.width
        img = img.resize((max_width, max(1, round(img.height * scale))), Image.Resampling.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.06)
    img = ImageEnhance.Sharpness(img).enhance(1.08)
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=78, optimize=True)
    return out.getvalue()


def pdf_to_images(pdf_bytes: bytes, max_pages: int = 10) -> list[Image.Image]:
    """Render PDF pages using pypdfium2 if available."""
    try:
        import pypdfium2 as pdfium
    except ImportError as exc:
        raise RuntimeError(
            "PDF support requires pypdfium2. Install it with: pip install pypdfium2"
        ) from exc

    pdf = pdfium.PdfDocument(pdf_bytes)
    pages = []
    for i in range(min(len(pdf), max_pages)):
        bitmap = pdf[i].render(scale=2.2)
        pages.append(bitmap.to_pil().convert("RGB"))
    return pages


def pil_to_png_bytes(img: Image.Image) -> bytes:
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()


# -----------------------------------------------------------------------------
# Optional docTR pass
# -----------------------------------------------------------------------------
@st.cache_resource(show_spinner="Loading docTR model (first run only)...")
def load_doctr_model():
    from doctr.models import ocr_predictor
    return ocr_predictor(det_arch="db_resnet50", reco_arch="parseq", pretrained=True)


def run_doctr(model, preprocessed: np.ndarray) -> str:
    rgb = cv2.cvtColor(preprocessed, cv2.COLOR_GRAY2RGB)
    result = model([rgb])
    exported = result.export()
    lines = []
    for page in exported.get("pages", []):
        for block in page.get("blocks", []):
            for line in block.get("lines", []):
                words = [w.get("value", "") for w in line.get("words", [])]
                lines.append(" ".join(x for x in words if x))
    return "\n".join(lines).strip()


# -----------------------------------------------------------------------------
# Dynamic prompts
# -----------------------------------------------------------------------------
APP_VERSION = "7.0-full-sheet-multipass"

DISCOVERY_AND_EXTRACTION_PROMPT = r"""
You are a multilingual document-understanding system specialized in attendance/absence registers.
This is PASS 1: PAGE STRUCTURE, HEADERS, METADATA, AND REGIONS.

The page may use ANY language, script, country, organization, school/company template, or text direction.
Preserve visible text exactly as written. Do not translate names or labels.

IMPORTANT: do NOT try to read all tiny handwritten attendance marks or all handwritten names in this pass.
Dedicated high-resolution passes will read those regions later. Your job here is to map the full sheet correctly.

Extract:
1. exact document title;
2. detected languages and text direction;
3. every visible metadata label/value anywhere above or around the table;
4. instruction/help text printed on the form, if present;
5. every identity column header in visible order (serial number, enrolment ID, name, sex, class, etc.);
6. every attendance-unit header in visible order (1..31, dates, weekdays, sessions, AM/PM, etc.);
7. every summary column header after the attendance area;
8. status/absence-code legend printed on the page;
9. tight normalized 0..1000 bounding boxes for the complete table and its three regions.

Return ONLY valid JSON:
{
  "is_attendance_sheet": true,
  "document": {
    "title": "exact visible title",
    "languages": ["English"],
    "text_direction": "ltr|rtl|mixed",
    "template_description": "short description",
    "document_notes": ""
  },
  "metadata": [
    {"label": "exact label", "value": "exact visible value"}
  ],
  "instructions": ["exact visible instruction line"],
  "identity_columns": [
    {"key": "field_001", "label": "exact visible header"}
  ],
  "attendance_columns": [
    {"key": "att_001", "label": "exact visible attendance header"}
  ],
  "summary_columns": [
    {"key": "summary_001", "label": "exact visible summary header"}
  ],
  "status_legend": {
    "P": "present",
    "A": "absent",
    "L": "late",
    "E": "excused",
    "H": "holiday/no session",
    "?": "unclear"
  },
  "table_bbox": [0, 0, 1000, 1000],
  "identity_grid_bbox": [0, 0, 300, 1000],
  "attendance_grid_bbox": [300, 0, 900, 1000],
  "summary_grid_bbox": [900, 0, 1000, 1000],
  "rows": []
}

Bounding-box rules:
- Coordinates are [left, top, right, bottom] on a 0..1000 scale relative to the ORIGINAL page image.
- table_bbox must include the column-header row and all visible participant/data rows, but not unrelated footer text.
- identity_grid_bbox must include ALL identity headers and the handwritten identity values beneath them.
- attendance_grid_bbox must include attendance headers and all attendance cells, excluding identity and summary columns.
- summary_grid_bbox must include summary headers and summary cells; if there are no summary columns, use null.
- It is better to include a few pixels of neighboring grid line than to cut off handwriting.

CRITICAL:
- Detect ALL column headers. Do not collapse 31 date/day columns into a single field.
- Do not invent participant rows in this pass. Leave rows as [].
- If a metadata value is handwritten, read it if clearly visible.
- If instruction text is too small, include only what is actually readable.
- Preserve the form's own legend/code meanings when visible.

Optional OCR hint below may be noisy; trust the image over OCR:
---
__OCR_HINT_PLACEHOLDER__
---
"""


def make_verification_prompt(data: dict) -> str:
    att_cols = data.get("attendance_columns", [])
    labels = [{"key": c.get("key", ""), "label": c.get("label", "")} for c in att_cols]
    row_count = len(data.get("rows", []))
    return f"""
You are doing a SECOND verification pass on an attendance register.
Do not assume any language, direction, date format, or template.

The first pass detected {row_count} participant rows and these attendance columns, in visible order:
{json.dumps(labels, ensure_ascii=False)}

Re-read ONLY the attendance grid cell-by-cell from the image.
Use row order to align participants. If a visible ID/name is easy to read, include it as row_hint.

Normalize statuses as follows:
P = present
A = absent
L = late
E = excused/authorized absence
H = whole-column holiday/non-working/no-session
? = visible but genuinely unclear
"" = truly blank

If another status is clearly defined by the sheet's own legend, preserve its short code.
Do not fill a blank simply because neighboring cells have marks.
Do not change column labels or reorder columns.

Return ONLY valid JSON:
{{
  "rows": [
    {{
      "row": 1,
      "row_hint": "visible ID/name or empty",
      "attendance": {{"att_001": "P"}}
    }}
  ]
}}

Include all {row_count} rows if they are visible, and include every listed attendance key for each row.
"""



# -----------------------------------------------------------------------------
# Precision attendance-grid helpers
# -----------------------------------------------------------------------------
def crop_from_normalized_bbox(img: Image.Image, bbox, pad_px: int = 3) -> Image.Image:
    """Crop [left, top, right, bottom] coordinates expressed on a 0..1000 scale."""
    try:
        l, t, r, b = [float(x) for x in bbox]
    except Exception:
        return img.copy()
    l = max(0, min(1000, l)); r = max(0, min(1000, r))
    t = max(0, min(1000, t)); b = max(0, min(1000, b))
    if r <= l or b <= t:
        return img.copy()
    x1 = max(0, int(round(l / 1000 * img.width)) - pad_px)
    y1 = max(0, int(round(t / 1000 * img.height)) - pad_px)
    x2 = min(img.width, int(round(r / 1000 * img.width)) + pad_px)
    y2 = min(img.height, int(round(b / 1000 * img.height)) + pad_px)
    return img.crop((x1, y1, x2, y2))


def prepare_precision_crop_bytes(img: Image.Image, target_width: int = 2200) -> bytes:
    """Upscale a SMALL grid crop so tiny handwriting is visible without enlarging the whole page."""
    img = img.convert("RGB")
    if img.width < target_width:
        scale = target_width / max(1, img.width)
        new_h = max(1, int(round(img.height * scale)))
        # Avoid pathological huge images for tall registers.
        if new_h > 2600:
            scale = 2600 / max(1, img.height)
            target_width = max(1, int(round(img.width * scale)))
            new_h = 2600
        img = img.resize((target_width, new_h), Image.Resampling.LANCZOS)
    img = ImageEnhance.Contrast(img).enhance(1.12)
    img = ImageEnhance.Sharpness(img).enhance(1.22)
    img = img.filter(ImageFilter.UnsharpMask(radius=1.0, percent=115, threshold=2))
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=90, optimize=True)
    return out.getvalue()


def make_chunk_prompt(row_count: int, columns: list[dict], start_index: int) -> str:
    labels = [str(c.get("label", "")) for c in columns]
    return f"""
You are reading ONLY a SMALL CROP of an attendance/absence grid.
The crop contains attendance columns {start_index + 1} through {start_index + len(columns)} of the full sheet.
Their visible labels, from LEFT TO RIGHT in this crop, are:
{json.dumps(labels, ensure_ascii=False)}

There are {row_count} participant rows. Read rows TOP TO BOTTOM.
Inspect EVERY cell independently at pixel level. Do not copy a pattern across a row.
Tiny dots, slashes, ticks, X marks, letters, and blanks are meaningful.

Normalize each cell ONLY when its meaning is clear from the form/legend:
P = present
A = absent
L = late
E = excused/authorized absence
H = holiday/non-working/no-session
? = a visible mark whose meaning is unclear
"" = truly blank

IMPORTANT:
- Do not assume later days are present because earlier days are present.
- Do not assume a blank cell has a status.
- Do not invent marks outside what is visible.
- Return exactly {row_count} arrays, one per visible participant row.
- Each row array must contain exactly {len(columns)} values.

Return ONLY valid compact JSON:
{{"rows":[["P","A"],["","P"]]}}
"""


def read_attendance_in_chunks(page: Image.Image, data: dict, model: str, timeout: int,
                              chunk_size: int = 6) -> dict:
    """Read a dense attendance grid in small horizontal chunks to avoid vision pattern hallucination."""
    rows = data.get("rows", [])
    cols = data.get("attendance_columns", [])
    if not rows or not cols:
        return data

    bbox = data.get("attendance_grid_bbox")
    if not isinstance(bbox, (list, tuple)) or len(bbox) != 4:
        # Conservative fallback: lower-middle portion of a typical attendance form.
        bbox = [300, 350, 930, 900]
    grid = crop_from_normalized_bbox(page, bbox, pad_px=4)

    n = len(cols)
    # The grid crop contains attendance columns only, which are usually equal-width.
    # Split by the exact number of detected headers, with a little overlap/padding.
    for start in range(0, n, chunk_size):
        end = min(n, start + chunk_size)
        x1 = int(round(start / n * grid.width))
        x2 = int(round(end / n * grid.width))
        pad = max(2, int(grid.width / n * 0.12))
        chunk_img = grid.crop((max(0, x1 - pad), 0, min(grid.width, x2 + pad), grid.height))
        chunk_bytes = prepare_precision_crop_bytes(chunk_img)
        prompt = make_chunk_prompt(len(rows), cols[start:end], start)
        result = ollama_vision_json(chunk_bytes, model, prompt, timeout, temperature=0.0,
                                    num_ctx=4096, num_predict=max(512, min(1400, len(rows) * (end-start) * 4 + 200)))
        out_rows = result.get("rows", [])
        for r_idx, target_row in enumerate(rows):
            if r_idx >= len(out_rows):
                continue
            values = out_rows[r_idx]
            if isinstance(values, dict):
                values = list(values.values())
            if not isinstance(values, list):
                continue
            att = target_row.setdefault("attendance", {})
            for local_i, col in enumerate(cols[start:end]):
                if local_i < len(values):
                    v = str(values[local_i] or "").strip().upper()
                    if v in {"P", "A", "L", "E", "H", "?", ""}:
                        att[col["key"]] = v
    return data


# -----------------------------------------------------------------------------
# Full-sheet multipass readers
# -----------------------------------------------------------------------------
def _fallback_identity_bbox(data: dict):
    table = data.get("table_bbox")
    att = data.get("attendance_grid_bbox")
    if isinstance(table, (list, tuple)) and len(table) == 4 and isinstance(att, (list, tuple)) and len(att) == 4:
        return [table[0], table[1], att[0], table[3]]
    return [0, 300, 360, 930]


def _fallback_summary_bbox(data: dict):
    table = data.get("table_bbox")
    att = data.get("attendance_grid_bbox")
    if isinstance(table, (list, tuple)) and len(table) == 4 and isinstance(att, (list, tuple)) and len(att) == 4:
        return [att[2], table[1], table[2], table[3]]
    return [900, 300, 1000, 930]


def make_identity_prompt(identity_columns: list[dict]) -> str:
    labels = [{"key": c.get("key", ""), "label": c.get("label", "")} for c in identity_columns]
    return f"""
You are reading a HIGH-RESOLUTION CROP containing the identity section of an attendance register.
It includes the identity column headers plus ALL visible participant rows beneath them.

The detected identity columns, in LEFT-TO-RIGHT visual order, are:
{json.dumps(labels, ensure_ascii=False)}

Your most important job is to extract EVERY FILLED participant row exactly as written.
Examples of identity fields may include serial number, enrolment number, student/employee full name, sex/gender, age, grade, class, ID, or other template-specific fields.

Rules:
- Preserve handwritten names EXACTLY as visible. Do not translate or modernize spelling.
- Return ALL filled participant rows from TOP TO BOTTOM.
- Do not stop after the first 2-5 rows. Inspect the entire crop down to the final filled row.
- Ignore completely unused blank rows at the bottom of a form.
- A row with a handwritten name/ID/sex is a participant row even if some identity fields are blank.
- If a field is genuinely unreadable, use "[unclear]" rather than inventing text.
- Do not extract attendance marks here.
- Keep the number and order of values exactly aligned to the listed identity columns.

Return ONLY compact valid JSON in this exact shape:
{{"rows":[["value for first identity column","value for second identity column"],["...","..."]]}}

Each returned row array MUST contain exactly {len(identity_columns)} values.
"""


def read_identity_rows(page: Image.Image, data: dict, model: str, timeout: int) -> dict:
    cols = data.get("identity_columns", [])
    if not cols:
        return data
    bbox = data.get("identity_grid_bbox")
    if not isinstance(bbox, (list, tuple)) or len(bbox) != 4:
        bbox = _fallback_identity_bbox(data)
    crop = crop_from_normalized_bbox(page, bbox, pad_px=8)
    crop_bytes = prepare_precision_crop_bytes(crop, target_width=2400)
    result = ollama_vision_json(
        crop_bytes, model, make_identity_prompt(cols), timeout,
        temperature=0.0, num_ctx=4096, num_predict=2200,
    )
    raw_rows = result.get("rows", [])
    rows = []
    for raw in raw_rows:
        if isinstance(raw, dict):
            vals = [str(raw.get(c.get("key", ""), "") or "") for c in cols]
        elif isinstance(raw, list):
            vals = [str(v or "") for v in raw]
        else:
            continue
        if len(vals) < len(cols):
            vals += [""] * (len(cols) - len(vals))
        vals = vals[:len(cols)]
        # Keep only genuinely filled participant rows.
        if not any(v.strip() for v in vals):
            continue
        rows.append({
            "identity": {c["key"]: vals[i] for i, c in enumerate(cols)},
            "attendance": {},
            "summary": {},
        })
    data["rows"] = rows
    normalize_structure(data)
    return data


def make_summary_prompt(summary_columns: list[dict], row_count: int) -> str:
    labels = [str(c.get("label", "")) for c in summary_columns]
    return f"""
You are reading a HIGH-RESOLUTION CROP containing ONLY summary columns from an attendance register.
The visible summary columns from LEFT TO RIGHT are:
{json.dumps(labels, ensure_ascii=False)}

There are {row_count} participant rows, ordered TOP TO BOTTOM.
Read each summary cell exactly as visible. Blank means blank. Do not calculate or infer totals.
Return exactly {row_count} row arrays and exactly {len(summary_columns)} values per row.

Return ONLY compact valid JSON:
{{"rows":[[""],["3"]]}}
"""


def read_summary_rows(page: Image.Image, data: dict, model: str, timeout: int) -> dict:
    cols = data.get("summary_columns", [])
    rows = data.get("rows", [])
    if not cols or not rows:
        return data
    bbox = data.get("summary_grid_bbox")
    if not isinstance(bbox, (list, tuple)) or len(bbox) != 4:
        bbox = _fallback_summary_bbox(data)
    crop = crop_from_normalized_bbox(page, bbox, pad_px=8)
    crop_bytes = prepare_precision_crop_bytes(crop, target_width=1200)
    result = ollama_vision_json(
        crop_bytes, model, make_summary_prompt(cols, len(rows)), timeout,
        temperature=0.0, num_ctx=3072, num_predict=max(600, min(1800, len(rows) * len(cols) * 8 + 200)),
    )
    raw_rows = result.get("rows", [])
    for r_idx, row in enumerate(rows):
        if r_idx >= len(raw_rows):
            break
        raw = raw_rows[r_idx]
        if isinstance(raw, dict):
            vals = [str(raw.get(c.get("key", ""), "") or "") for c in cols]
        elif isinstance(raw, list):
            vals = [str(v or "") for v in raw]
        else:
            continue
        if len(vals) < len(cols):
            vals += [""] * (len(cols) - len(vals))
        for i, col in enumerate(cols):
            row.setdefault("summary", {})[col["key"]] = vals[i] if i < len(vals) else ""
    return data

# -----------------------------------------------------------------------------
# Ollama helpers
# -----------------------------------------------------------------------------
def list_ollama_models() -> list[str]:
    try:
        resp = requests.get(OLLAMA_TAGS_URL, timeout=5)
        resp.raise_for_status()
        return [m["name"] for m in resp.json().get("models", [])]
    except Exception:
        return []


def _extract_balanced_json(text: str) -> str:
    """Return the first balanced top-level JSON object from model output."""
    start = text.find("{")
    if start < 0:
        return text.strip()
    depth = 0
    in_string = False
    escaped = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_string:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            continue
        if ch == '"':
            in_string = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return text[start:].strip()


def _repair_common_json_errors(text: str) -> str:
    """Repair common small JSON mistakes produced by vision LLMs."""
    s = text.strip()
    # Normalize typographic punctuation that sometimes leaks into JSON.
    s = s.replace("“", '"').replace("”", '"').replace("„", '"')
    s = s.replace("’", "'").replace("‘", "'")
    # Remove JavaScript-style comments and trailing commas.
    s = re.sub(r"/\*.*?\*/", "", s, flags=re.DOTALL)
    s = re.sub(r"(?m)^\s*//.*$", "", s)
    s = re.sub(r",\s*([}\]])", r"\1", s)
    # Insert commas between adjacent JSON members/values when the model forgot one.
    # Examples:  "name": "Ali"  "age": "10"
    #            } {   or   ] "next_key":
    s = re.sub(r'("(?:[^"\\]|\\.)*"|true|false|null|-?\d+(?:\.\d+)?)\s*(?="[^"\n]+"\s*:)', r'\1,', s)
    s = re.sub(r'([}\]])\s*(?=[{\[])', r'\1,', s)
    s = re.sub(r'([}\]])\s*(?="[^"\n]+"\s*:)', r'\1,', s)
    return s


def parse_json_response(raw: str) -> dict:
    cleaned = raw.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    cleaned = _extract_balanced_json(cleaned)

    attempts = [cleaned, _repair_common_json_errors(cleaned)]
    last_error = None
    for candidate in attempts:
        try:
            return json.loads(candidate)
        except json.JSONDecodeError as e:
            last_error = e

    # Optional stronger repair if the tiny `json-repair` package is installed.
    try:
        from json_repair import repair_json
        repaired = repair_json(cleaned)
        if isinstance(repaired, dict):
            return repaired
        return json.loads(repaired)
    except Exception:
        pass

    if last_error is not None:
        line = last_error.lineno
        col = last_error.colno
        lines = cleaned.splitlines()
        context = "\n".join(lines[max(0, line-2): min(len(lines), line+1)])
        raise ValueError(
            f"Vision model returned malformed JSON at line {line}, column {col}: {last_error.msg}.\n"
            f"Nearby model output:\n{context}\n\n"
            "The page was read, but the model made a JSON punctuation error. "
            "Try again once; if it repeats, install `json-repair` with: pip install json-repair"
        )
    raise ValueError("Vision model returned empty or invalid JSON.")


def ollama_vision_json(image_bytes: bytes, model: str, prompt: str, timeout: int, temperature: float = 0.0,
                       num_ctx: int = 4096, num_predict: int = 1536) -> dict:
    payload = {
        "model": model,
        "prompt": prompt,
        "images": [base64.b64encode(image_bytes).decode("utf-8")],
        "stream": False,
        "options": {
            "temperature": temperature,
            "num_ctx": num_ctx,
            "num_predict": num_predict,
            "num_thread": 8,
        },
    }
    resp = requests.post(OLLAMA_GENERATE_URL, json=payload, timeout=timeout)
    if resp.status_code == 400:
        try:
            detail = resp.json().get("error", resp.text)
        except ValueError:
            detail = resp.text
        raise ValueError(
            f"Ollama rejected the image request: {detail}\n"
            "Choose a multimodal model such as qwen2.5vl, qwen3-vl, llama3.2-vision, minicpm-v, or llava."
        )
    resp.raise_for_status()
    return parse_json_response(resp.json().get("response", ""))


def extract_sheet_data(image_bytes: bytes, model: str, timeout: int, ocr_text: str = "") -> dict:
    prompt = DISCOVERY_AND_EXTRACTION_PROMPT.replace("__OCR_HINT_PLACEHOLDER__", ocr_text or "(no OCR hint available)")
    data = ollama_vision_json(image_bytes, model, prompt, timeout, temperature=0.0, num_ctx=6144, num_predict=2600)
    normalize_structure(data)
    return data


def verify_attendance(image_bytes: bytes, model: str, timeout: int, data: dict) -> dict:
    return ollama_vision_json(image_bytes, model, make_verification_prompt(data), timeout, temperature=0.0)


# -----------------------------------------------------------------------------
# Dynamic schema normalization / merge
# -----------------------------------------------------------------------------
def normalize_structure(data: dict) -> None:
    data.setdefault("document", {})
    data["document"].setdefault("title", "Attendance Sheet")
    data["document"].setdefault("languages", [])
    data["document"].setdefault("text_direction", "ltr")
    data["document"].setdefault("template_description", "")
    data["document"].setdefault("document_notes", "")
    data.setdefault("metadata", [])
    data.setdefault("identity_columns", [])
    data.setdefault("attendance_columns", [])
    data.setdefault("summary_columns", [])
    data.setdefault("status_legend", {})
    data.setdefault("rows", [])
    data.setdefault("table_bbox", None)
    data.setdefault("identity_grid_bbox", None)
    data.setdefault("attendance_grid_bbox", None)
    data.setdefault("summary_grid_bbox", None)
    data.setdefault("instructions", [])

    # Repair missing/duplicate keys without changing visible labels.
    def fix_keys(columns, prefix):
        used = set()
        for idx, col in enumerate(columns, 1):
            key = str(col.get("key", "")).strip() or f"{prefix}_{idx:03d}"
            base = re.sub(r"[^A-Za-z0-9_]+", "_", key).strip("_") or f"{prefix}_{idx:03d}"
            key = base
            n = 2
            while key in used:
                key = f"{base}_{n}"
                n += 1
            col["key"] = key
            col.setdefault("label", key)
            used.add(key)

    fix_keys(data["identity_columns"], "field")
    fix_keys(data["attendance_columns"], "att")
    fix_keys(data["summary_columns"], "summary")

    identity_keys = [c["key"] for c in data["identity_columns"]]
    attendance_keys = [c["key"] for c in data["attendance_columns"]]
    summary_keys = [c["key"] for c in data["summary_columns"]]

    for row in data["rows"]:
        identity = row.setdefault("identity", {})
        attendance = row.setdefault("attendance", {})
        summary = row.setdefault("summary", {})
        for key in identity_keys:
            identity.setdefault(key, "")
        for key in attendance_keys:
            attendance.setdefault(key, "")
        for key in summary_keys:
            summary.setdefault(key, "")


def merge_attendance_fill_blanks(data: dict, verified: dict) -> dict:
    rows = data.get("rows", [])
    vrows = verified.get("rows", [])
    valid_keys = {c.get("key") for c in data.get("attendance_columns", [])}

    for idx, row in enumerate(rows):
        if idx >= len(vrows):
            break
        target = row.setdefault("attendance", {})
        candidate = vrows[idx].get("attendance", {}) or {}
        for key in valid_keys:
            current = str(target.get(key, "") or "").strip()
            proposed = str(candidate.get(key, "") or "").strip()
            if not current and proposed:
                target[key] = proposed
    return data


# -----------------------------------------------------------------------------
# DataFrame conversion
# -----------------------------------------------------------------------------
def data_to_dataframe(data: dict) -> pd.DataFrame:
    id_cols = data.get("identity_columns", [])
    att_cols = data.get("attendance_columns", [])
    sum_cols = data.get("summary_columns", [])

    rows_out = []
    for row in data.get("rows", []):
        flat = {}
        for col in id_cols:
            flat[col["label"]] = row.get("identity", {}).get(col["key"], "")
        for col in att_cols:
            flat[col["label"]] = row.get("attendance", {}).get(col["key"], "")
        for col in sum_cols:
            flat[col["label"]] = row.get("summary", {}).get(col["key"], "")
        rows_out.append(flat)

    columns = [c["label"] for c in id_cols + att_cols + sum_cols]
    return pd.DataFrame(rows_out, columns=columns)


def dataframe_to_data(df: pd.DataFrame, data: dict, metadata_pairs: list[dict]) -> dict:
    out = json.loads(json.dumps(data, ensure_ascii=False))
    out["metadata"] = metadata_pairs
    out_rows = []

    for _, rec in df.fillna("").iterrows():
        row = {"identity": {}, "attendance": {}, "summary": {}}
        for col in out.get("identity_columns", []):
            row["identity"][col["key"]] = str(rec.get(col["label"], ""))
        for col in out.get("attendance_columns", []):
            row["attendance"][col["key"]] = str(rec.get(col["label"], ""))
        for col in out.get("summary_columns", []):
            row["summary"][col["key"]] = str(rec.get(col["label"], ""))
        out_rows.append(row)

    out["rows"] = out_rows
    return out


# -----------------------------------------------------------------------------
# Generic Excel builder
# -----------------------------------------------------------------------------
def build_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    direction = str(data.get("document", {}).get("text_direction", "ltr")).lower()
    ws.sheet_view.rightToLeft = direction == "rtl"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    meta_fill = PatternFill("solid", fgColor="E2F0D9")
    holiday_fill = PatternFill("solid", fgColor="F2F2F2")
    flag_fill = PatternFill("solid", fgColor="FFF2CC")

    normal = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=14, bold=True)

    id_cols = data.get("identity_columns", [])
    att_cols = data.get("attendance_columns", [])
    sum_cols = data.get("summary_columns", [])
    all_cols = id_cols + att_cols + sum_cols
    total_cols = max(1, len(all_cols))

    title = data.get("document", {}).get("title", "Attendance Sheet") or "Attendance Sheet"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = title_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center")

    row_idx = 3
    for item in data.get("metadata", []):
        label = str(item.get("label", ""))
        value = str(item.get("value", ""))
        ws.cell(row_idx, 1, label).font = bold
        ws.cell(row_idx, 1).fill = meta_fill
        if total_cols > 1:
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=total_cols)
        ws.cell(row_idx, 2 if total_cols > 1 else 1, value).font = normal
        row_idx += 1

    if data.get("metadata"):
        row_idx += 1

    instructions = data.get("instructions", []) or []
    if instructions:
        ws.cell(row_idx, 1, "Instructions").font = bold
        row_idx += 1
        for line in instructions:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
            ws.cell(row_idx, 1, str(line)).alignment = Alignment(wrap_text=True, vertical="top")
            row_idx += 1
        row_idx += 1

    header_row = row_idx
    for col_idx, col in enumerate(all_cols, 1):
        cell = ws.cell(header_row, col_idx, col.get("label", col.get("key", "")))
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(data.get("rows", []), header_row + 1):
        cidx = 1
        for col in id_cols:
            cell = ws.cell(r_idx, cidx, row.get("identity", {}).get(col["key"], ""))
            cell.font = normal
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if direction == "rtl" else "left", vertical="center")
            cidx += 1

        for col in att_cols:
            mark = str(row.get("attendance", {}).get(col["key"], "") or "")
            display = {"P": "✓", "A": "X", "H": "—"}.get(mark, mark)
            cell = ws.cell(r_idx, cidx, display)
            cell.font = normal
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if mark == "H":
                cell.fill = holiday_fill
            elif mark == "?":
                cell.fill = flag_fill
            cidx += 1

        for col in sum_cols:
            cell = ws.cell(r_idx, cidx, row.get("summary", {}).get(col["key"], ""))
            cell.font = normal
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cidx += 1

    # Column widths
    for idx, col in enumerate(id_cols, 1):
        label = str(col.get("label", ""))
        ws.column_dimensions[get_column_letter(idx)].width = max(10, min(28, len(label) + 5))

    att_start = len(id_cols) + 1
    for i, col in enumerate(att_cols):
        label = str(col.get("label", ""))
        ws.column_dimensions[get_column_letter(att_start + i)].width = max(4, min(12, len(label) + 2))

    sum_start = len(id_cols) + len(att_cols) + 1
    for i, col in enumerate(sum_cols):
        label = str(col.get("label", ""))
        ws.column_dimensions[get_column_letter(sum_start + i)].width = max(10, min(24, len(label) + 4))

    ws.freeze_panes = ws.cell(header_row + 1, max(1, len(id_cols) + 1))

    # Legend / detected information
    legend_row = header_row + len(data.get("rows", [])) + 2
    ws.cell(legend_row, 1, "Detected languages").font = bold
    ws.cell(legend_row, 2 if total_cols > 1 else 1, ", ".join(data.get("document", {}).get("languages", [])))
    legend_row += 1
    ws.cell(legend_row, 1, "Status legend").font = bold
    legend_text = " | ".join(f"{k} = {v}" for k, v in data.get("status_legend", {}).items())
    if total_cols > 1:
        ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=total_cols)
        ws.cell(legend_row, 2, legend_text)
    else:
        ws.cell(legend_row, 1, legend_text)

    # Print/PDF layout: fit arbitrary-width templates to one page wide.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if total_cols > 8 else ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # may use multiple pages vertically for very long registers
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.20
    ws.page_margins.right = 0.20
    ws.page_margins.top = 0.30
    ws.page_margins.bottom = 0.30
    ws.print_area = f"A1:{get_column_letter(total_cols)}{legend_row}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------------------------------------------------------
# PDF export from generated workbook
# -----------------------------------------------------------------------------
def excel_to_pdf(excel_bytes: bytes) -> bytes:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        for candidate in [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]:
            if Path(candidate).exists():
                soffice = candidate
                break
    if not soffice:
        raise RuntimeError(
            "LibreOffice was not found. Install LibreOffice or add its program folder to PATH to enable PDF export."
        )

    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        xlsx_path = tmp_path / "attendance.xlsx"
        xlsx_path.write_bytes(excel_bytes)
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(tmp_path), str(xlsx_path)],
            capture_output=True,
            text=True,
            timeout=120,
        )
        pdf_path = tmp_path / "attendance.pdf"
        if not pdf_path.exists():
            raise RuntimeError(f"LibreOffice conversion failed:\n{result.stdout}\n{result.stderr}")
        return pdf_path.read_bytes()


# -----------------------------------------------------------------------------
# Streamlit app
# -----------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="Universal Attendance Extractor", layout="wide")
    st.title("📋 Universal Attendance Sheet Extractor")
    st.caption(f"App version: {APP_VERSION}")
    st.caption(
        "Precision mode: detects any attendance template first, then reads dense attendance cells in small high-resolution chunks instead of hallucinating long patterns."
    )

    if "extracted_pages" not in st.session_state:
        st.session_state.extracted_pages = []

    with st.sidebar:
        st.header("Settings")
        models = list_ollama_models()
        if models:
            preferred = next((i for i, m in enumerate(models) if "3b" in m.lower() and ("vision" in m.lower() or "vl" in m.lower())), next((i for i, m in enumerate(models) if "vision" in m.lower() or "vl" in m.lower()), 0))
            model_name = st.selectbox("Ollama vision model", models, index=preferred)
        else:
            model_name = st.text_input("Ollama vision model", DEFAULT_MODEL)
            st.caption("Ollama was not detected at localhost:11434.")

        timeout = st.slider("Extraction timeout (seconds)", 60, 1800, 1800, step=60)
        precision_grid = st.checkbox("Precision chunked grid reading", value=True, help="Reads every attendance cell from high-resolution grid crops after a separate high-resolution name/identity pass.")
        chunk_size = st.slider("Attendance columns per vision pass", 3, 10, 6, step=1)
        verify_grid = st.checkbox("Extra blank-cell verification pass", value=False)
        use_ocr = st.checkbox("Use docTR as an optional OCR hint", value=False)
        st.caption("The page is read in multiple passes: structure/metadata, full identity/name rows, attendance chunks, then summary columns. This is designed to capture the ENTIRE filled sheet rather than only recreating the template.")
        block_size = st.slider("Threshold block size", 5, 51, 25, step=2, disabled=not use_ocr)
        c_value = st.slider("Threshold offset", 0, 30, 15, disabled=not use_ocr)

    uploaded = st.file_uploader(
        "Upload an attendance sheet",
        type=["png", "jpg", "jpeg", "webp", "pdf"],
        help="The layout and language are detected automatically.",
    )
    if uploaded is None:
        st.info("Upload an image or PDF attendance register to begin.")
        return

    suffix = Path(uploaded.name).suffix.lower()
    raw = uploaded.getvalue()

    try:
        if suffix == ".pdf":
            pages = pdf_to_images(raw)
        else:
            pages = [Image.open(io.BytesIO(raw)).convert("RGB")]
    except Exception as exc:
        st.error(f"Could not open this file: {exc}")
        return

    st.write(f"Detected {len(pages)} page(s).")
    preview_cols = st.columns(min(3, len(pages)))
    for i, page in enumerate(pages[:3]):
        with preview_cols[i % len(preview_cols)]:
            st.image(page, caption=f"Page {i + 1}", width="stretch")

    if st.button("Extract attendance", type="primary"):
        extracted_pages = []
        doctr_model = None
        if use_ocr:
            try:
                doctr_model = load_doctr_model()
            except Exception as exc:
                st.warning(f"docTR could not be loaded, so extraction will continue with vision only: {exc}")

        for page_index, page in enumerate(pages, 1):
            with st.spinner(f"Reading page {page_index}/{len(pages)} with {model_name}..."):
                try:
                    page_bytes = pil_to_png_bytes(page)
                    # Structure pass: keep more detail than the old 1280px turbo image,
                    # but do not ask it to read tiny attendance marks.
                    vision_bytes = prepare_vision_image_bytes(page_bytes, max_width=1800)
                    ocr_text = ""
                    if doctr_model is not None:
                        preprocessed = preprocess_image(page, block_size, c_value)
                        ocr_text = run_doctr(doctr_model, preprocessed)

                    data = extract_sheet_data(vision_bytes, model_name, timeout, ocr_text)

                    # PASS 2: names / IDs / sex / identity values at high resolution.
                    with st.spinner(f"Page {page_index}: reading every participant name and identity row..."):
                        data = read_identity_rows(page, data, model_name, timeout)

                    # PASS 3: every attendance cell in small horizontal chunks.
                    if precision_grid and data.get("attendance_columns") and data.get("rows"):
                        with st.spinner(f"Page {page_index}: reading {len(data.get('attendance_columns', []))} attendance columns for {len(data.get('rows', []))} participant rows..."):
                            data = read_attendance_in_chunks(page, data, model_name, timeout, chunk_size=chunk_size)

                    # PASS 4: totals / summary columns.
                    if data.get("summary_columns") and data.get("rows"):
                        with st.spinner(f"Page {page_index}: reading summary/total columns..."):
                            data = read_summary_rows(page, data, model_name, timeout)

                    if not data.get("is_attendance_sheet", True):
                        st.warning(
                            f"Page {page_index} may not be an attendance sheet: "
                            f"{data.get('document', {}).get('document_notes', '')}"
                        )

                    if verify_grid and data.get("attendance_columns") and data.get("rows"):
                        verified = verify_attendance(vision_bytes, model_name, timeout, data)
                        data = merge_attendance_fill_blanks(data, verified)

                    extracted_pages.append(data)
                except requests.exceptions.Timeout:
                    st.error(f"Page {page_index} timed out after {timeout} seconds.")
                except requests.exceptions.RequestException as exc:
                    st.error(f"Could not reach Ollama while reading page {page_index}: {exc}")
                except Exception as exc:
                    st.error(f"Page {page_index} failed: {type(exc).__name__}: {exc}")
                    with st.expander(f"Technical traceback for page {page_index}"):
                        st.code(traceback.format_exc())

        st.session_state.extracted_pages = extracted_pages
        if extracted_pages:
            st.success("Extraction complete. Review and edit the detected tables below.")

    extracted_pages = st.session_state.extracted_pages
    if not extracted_pages:
        return

    for page_index, data in enumerate(extracted_pages, 1):
        st.divider()
        doc = data.get("document", {})
        st.subheader(f"Page {page_index}: {doc.get('title', 'Attendance Sheet')}")
        st.caption(
            f"Language(s): {', '.join(doc.get('languages', [])) or 'Unknown'} | "
            f"Direction: {doc.get('text_direction', 'unknown')} | "
            f"Template: {doc.get('template_description', '')}"
        )

        # Dynamic metadata editor
        metadata = data.get("metadata", [])
        edited_meta = []
        if metadata:
            st.markdown("**Detected metadata**")
            for i, item in enumerate(metadata):
                c1, c2 = st.columns([1, 2])
                with c1:
                    label = st.text_input(
                        "Metadata label",
                        value=str(item.get("label", "")),
                        key=f"meta_label_{page_index}_{i}",
                    )
                with c2:
                    value = st.text_input(
                        "Metadata value",
                        value=str(item.get("value", "")),
                        key=f"meta_value_{page_index}_{i}",
                    )
                edited_meta.append({"label": label, "value": value})
        else:
            edited_meta = []

        st.markdown("**Attendance table — editable**")
        st.caption("Precision result. P = present, A = absent, L = late, E = excused, H = holiday/no-session, ? = unclear. Blank means the cell was read as blank.")
        df = data_to_dataframe(data)
        edited_df = st.data_editor(
            df,
            width="stretch",
            num_rows="dynamic",
            height=min(650, 120 + max(1, len(df)) * 35),
            key=f"editor_{page_index}",
        )

        final_data = dataframe_to_data(edited_df, data, edited_meta)
        excel_bytes = build_excel(final_data)

        safe_base = re.sub(r"[^A-Za-z0-9_-]+", "_", str(doc.get("title", "attendance"))).strip("_") or "attendance"
        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                "⬇️ Download Excel",
                data=excel_bytes,
                file_name=f"{safe_base}_page_{page_index}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"xlsx_{page_index}",
            )
        with c2:
            if st.button("Prepare PDF", key=f"pdf_button_{page_index}"):
                try:
                    pdf_bytes = excel_to_pdf(excel_bytes)
                    st.download_button(
                        "⬇️ Download PDF",
                        data=pdf_bytes,
                        file_name=f"{safe_base}_page_{page_index}.pdf",
                        mime="application/pdf",
                        key=f"pdf_download_{page_index}",
                    )
                except RuntimeError as exc:
                    st.error(str(exc))

        with st.expander(f"Raw JSON — page {page_index}"):
            st.json(final_data)


if __name__ == "__main__":
    main()