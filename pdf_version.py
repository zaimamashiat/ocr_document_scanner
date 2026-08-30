import inspect
import io
import json
import math
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import pypdfium2 as pdfium
import requests
import streamlit as st
from PIL import Image

from paddleocr import PaddleOCR

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


# ============================================================
# CONFIG
# ============================================================

OLLAMA_URL = "http://localhost:11434/api/chat"
OLLAMA_MODEL = "gemma3:12b"

# PaddleOCR language codes. PaddleOCR bundles several Latin languages
# (French/German/Spanish/etc.) under a shared "latin" model, and several
# others (Malay/Indonesian/Swahili/etc.) under a few other shared packs.
# See https://github.com/PaddlePaddle/PaddleOCR for the full list — this
# covers the ones people actually digitize attendance sheets in.
LANGUAGE_OPTIONS = {
    "English": "en",
    "Arabic": "arabic",
    "Chinese (Simplified)": "ch",
    "Chinese (Traditional)": "chinese_cht",
    "French": "french",
    "German": "german",
    "Spanish / Portuguese / Italian (Latin script)": "latin",
    "Korean": "korean",
    "Japanese": "japan",
    "Russian / Ukrainian (Cyrillic)": "cyrillic",
    "Hindi / Marathi / Nepali (Devanagari)": "devanagari",
    "Tamil": "ta",
    "Telugu": "te",
    "Kannada": "ka",
    "Vietnamese": "vi",
}

st.set_page_config(
    page_title="Document Digitizer",
    page_icon="📋",
    layout="wide",
)

st.title("📋 Attendance / Ledger Digitizer")
st.caption("PaddleOCR + OpenCV + Ollama → Excel / CSV / PDF")


def streamlit_supports_param(func, param_name):
    try:
        return param_name in inspect.signature(func).parameters
    except (TypeError, ValueError):
        return False


def st_image_compat(*args, use_container_width=True, **kwargs):
    if use_container_width and streamlit_supports_param(st.image, "use_container_width"):
        kwargs["use_container_width"] = True
    return st.image(*args, **kwargs)


def st_button_compat(*args, use_container_width=True, **kwargs):
    if use_container_width and streamlit_supports_param(st.button, "use_container_width"):
        kwargs["use_container_width"] = True
    return st.button(*args, **kwargs)


def st_data_editor_compat(*args, use_container_width=True, **kwargs):
    if use_container_width and streamlit_supports_param(st.data_editor, "use_container_width"):
        kwargs["use_container_width"] = True
    return st.data_editor(*args, **kwargs)


def st_download_button_compat(*args, use_container_width=True, **kwargs):
    if use_container_width and streamlit_supports_param(st.download_button, "use_container_width"):
        kwargs["use_container_width"] = True
    return st.download_button(*args, **kwargs)


# ============================================================
# OCR
# ============================================================

@st.cache_resource(show_spinner="Loading PaddleOCR model (first run downloads weights)...")
def get_ocr(lang: str = "en"):
    """
    Cached per-language so switching languages in the sidebar loads a
    fresh model instead of silently reusing the wrong one.

    PaddleOCR's constructor kwargs have changed across major versions
    (use_angle_cls -> use_textline_orientation, show_log removed, etc.).
    Try the modern signature first and fall back to the older one so this
    keeps working regardless of which PaddleOCR version is installed.
    """
    try:
        return PaddleOCR(
            use_textline_orientation=True,
            lang=lang,
        )
    except TypeError:
        return PaddleOCR(
            use_angle_cls=True,
            lang=lang,
            show_log=False,
        )


def _flatten_paddle_result(result):
    """
    PaddleOCR's return shape has varied across versions:
      - list of [box, (text, score)] per detected line (det=True)
      - list of (text, score) tuples (det=False, recognition-only)
      - newer versions wrap everything in a dict-like OCRResult object
        with 'rec_texts' / 'rec_scores' / 'rec_polys' keys.
    Normalize all of these into a flat list of (text, score, box|None).
    """
    out = []

    if result is None:
        return out

    # Newer (PaddleX-based) OCRResult: dict-like with rec_texts/rec_scores
    if hasattr(result, "get") or isinstance(result, dict):
        try:
            texts = result.get("rec_texts", []) or []
            scores = result.get("rec_scores", []) or []
            polys = result.get("rec_polys", result.get("dt_polys", [])) or []
            for i, text in enumerate(texts):
                score = scores[i] if i < len(scores) else 0.0
                box = polys[i] if i < len(polys) else None
                out.append((str(text).strip(), float(score), box))
            return out
        except Exception:
            pass

    # Older list-based formats
    for page in result:
        if page is None:
            continue
        # det=False recognition-only single result: page == (text, score)
        if (
            isinstance(page, (list, tuple))
            and len(page) == 2
            and isinstance(page[0], str)
        ):
            text, score = page
            out.append((str(text).strip(), float(score), None))
            continue
        # det=True full pipeline: page is a list of [box, (text, score)]
        if isinstance(page, (list, tuple)):
            for line in page:
                try:
                    box, (text, score) = line
                    out.append((str(text).strip(), float(score), box))
                except Exception:
                    continue

    return out


def ocr_words(model, img):
    """Full-page OCR (detection + recognition) -> [{text, confidence, box}]

    Kept for cases where you need to OCR a *whole page* with unknown
    layout. For pre-cropped cells with known boundaries, use
    reco_batch() instead — it skips detection entirely and is much
    faster.
    """

    if img is None or img.size == 0:
        return []

    if img.ndim == 2:
        img = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)

    img = img.astype(np.uint8)

    try:
        result = model.ocr(img, cls=True)
    except TypeError:
        result = model.ocr(img)

    output = []

    for text, confidence, box in _flatten_paddle_result(result):
        if not text:
            continue

        if box is not None:
            xs = [p[0] for p in box]
            ys = [p[1] for p in box]
            bbox = [int(min(xs)), int(min(ys)), int(max(xs)), int(max(ys))]
        else:
            bbox = [0, 0, img.shape[1], img.shape[0]]

        output.append({
            "text": text,
            "confidence": float(confidence),
            "box": bbox,
        })

    return output


def reco_batch(model, crops):
    """
    Recognition-only OCR on a list of already-cropped cell images.

    We already know exactly where each cell is (we cropped it
    ourselves), so detection (the "det" step) is skipped entirely via
    det=False — PaddleOCR treats the whole crop as one text line and
    goes straight to recognition, which is far cheaper than running
    full detection+recognition per cell.

    Returns a list of strings, same length and order as `crops`.
    Empty/invalid crops map to "".
    """

    results = []

    for c in crops:

        if c is None or c.size == 0:
            results.append("")
            continue

        img = (
            cv2.cvtColor(c, cv2.COLOR_GRAY2RGB)
            if c.ndim == 2
            else c
        )

        img = cv2.resize(
            img,
            None,
            fx=2,
            fy=2,
            interpolation=cv2.INTER_CUBIC,
        ).astype(np.uint8)

        text = ""

        try:
            try:
                out = model.ocr(img, det=False, cls=False)
            except TypeError:
                out = model.ocr(img, det=False)

            flat = _flatten_paddle_result(out)
            if flat:
                text = flat[0][0]
        except Exception:
            text = ""

        results.append(text)

    return results


# ============================================================
# FILE READING
# ============================================================

def read_file(file):
    data = file.getvalue()

    if Path(file.name).suffix.lower() == ".pdf":

        pdf = pdfium.PdfDocument(data)

        return [
            np.array(
                pdf[i]
                .render(scale=3)
                .to_pil()
                .convert("RGB")
            )
            for i in range(len(pdf))
        ]

    return [
        np.array(
            Image.open(io.BytesIO(data)).convert("RGB")
        )
    ]


# ============================================================
# PREPROCESSING
# ============================================================

def deskew(img):
    gray = cv2.cvtColor(
        img,
        cv2.COLOR_RGB2GRAY,
    )

    edges = cv2.Canny(gray, 50, 150)

    lines = cv2.HoughLinesP(
        edges,
        1,
        np.pi / 180,
        100,
        minLineLength=max(100, img.shape[1] // 4),
        maxLineGap=20,
    )

    if lines is None:
        return img

    angles = []

    # FIXES numpy.int32 unpack error
    for line in lines:

        values = np.asarray(line).reshape(-1)

        if values.size < 4:
            continue

        x1, y1, x2, y2 = map(
            int,
            values[:4],
        )

        angle = math.degrees(
            math.atan2(
                y2 - y1,
                x2 - x1,
            )
        )

        if abs(angle) < 10:
            angles.append(angle)

    if not angles:
        return img

    angle = float(np.median(angles))

    if abs(angle) < 0.1:
        return img

    h, w = img.shape[:2]

    matrix = cv2.getRotationMatrix2D(
        (w / 2, h / 2),
        angle,
        1,
    )

    return cv2.warpAffine(
        img,
        matrix,
        (w, h),
        borderMode=cv2.BORDER_REPLICATE,
    )


def preprocess(img):
    # Limit huge scans
    h, w = img.shape[:2]

    if w > 3500:
        scale = 3500 / w

        img = cv2.resize(
            img,
            None,
            fx=scale,
            fy=scale,
            interpolation=cv2.INTER_AREA,
        )

    img = deskew(img)

    gray = cv2.cvtColor(
        img,
        cv2.COLOR_RGB2GRAY,
    )

    gray = cv2.createCLAHE(
        2,
        (8, 8),
    ).apply(gray)

    return img, gray


# ============================================================
# TABLE LINES
# ============================================================

def group(values, tolerance=5):
    values = sorted(values)

    if not values:
        return []

    groups = [[values[0]]]

    for value in values[1:]:

        if abs(
            value - np.mean(groups[-1])
        ) <= tolerance:
            groups[-1].append(value)

        else:
            groups.append([value])

    return [
        int(np.mean(x))
        for x in groups
    ]


def detect_grid(img, gray):
    binary = cv2.adaptiveThreshold(
        gray,
        255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        31,
        13,
    )

    # Ruling lines on scanned attendance sheets are often thin red
    # ink, which loses most of its contrast once flattened to
    # grayscale (and can vanish further after CLAHE). That was
    # causing every other horizontal row line to go undetected,
    # which in turn made student_rows() merge two real rows into
    # one oversized crop -> garbled, "two names smashed together"
    # OCR output. Build a second mask straight from the color image
    # that specifically picks up reddish pixels, and OR it into the
    # structural mask so faint red lines still count.
    r, g, b = cv2.split(img.astype(np.int16))

    red_mask = (
        (r - np.maximum(g, b)) > 20
    ).astype(np.uint8) * 255

    binary = cv2.bitwise_or(binary, red_mask)

    h, w = gray.shape

    hk = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (max(20, w // 40), 1),
    )

    vk = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (1, max(20, h // 40)),
    )

    horizontal = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        hk,
    )

    vertical = cv2.morphologyEx(
        binary,
        cv2.MORPH_OPEN,
        vk,
    )

    hp = np.sum(horizontal > 0, axis=1)
    vp = np.sum(vertical > 0, axis=0)

    ys = group(
        np.where(hp > w * 0.12)[0].tolist()
    )

    xs = group(
        np.where(vp > h * 0.12)[0].tolist()
    )

    return xs, ys


# ============================================================
# ATTENDANCE STRUCTURE
# ============================================================

def find_day_columns(xs):
    """
    Find the longest sequence of narrow,
    approximately equal-width columns.
    """

    if len(xs) < 15:
        return None

    gaps = np.diff(xs)

    small = sorted(
        g for g in gaps
        if g > 3
    )

    if not small:
        return None

    width = np.median(
        small[:max(5, len(small) // 2)]
    )

    runs = []

    for start in range(len(xs) - 1):

        run = [xs[start]]

        for i in range(start, len(xs) - 1):

            gap = xs[i + 1] - xs[i]

            if width * 0.5 <= gap <= width * 1.7:
                run.append(xs[i + 1])

            else:
                break

        if len(run) >= 15:
            runs.append(run)

    if not runs:
        return None

    days = max(runs, key=len)

    # Attendance has 31 days = 32 boundaries
    return days[:32]


def recover_missing_lines(ys, min_gap=8):
    """
    Safety net for row lines detect_grid() still missed (e.g. an
    especially faint or broken rule line). If a gap between two
    detected lines is close to an integer multiple of the smallest
    plausible row height, it's almost certainly hiding one or more
    undetected lines in between — insert synthetic split points at
    even spacing so downstream cropping doesn't grab two students'
    rows as one (the cause of merged/garbled OCR text).
    """

    ys = sorted(set(int(y) for y in ys))

    if len(ys) < 3:
        return ys

    gaps = np.diff(ys)

    plausible = gaps[gaps > min_gap]

    if plausible.size == 0:
        return ys

    unit = float(np.min(plausible))

    if unit <= 0:
        return ys

    filled = [ys[0]]

    for y1, y2 in zip(ys[:-1], ys[1:]):

        gap = y2 - y1
        n = max(1, int(round(gap / unit)))

        for k in range(1, n):
            filled.append(int(round(y1 + k * gap / n)))

        filled.append(y2)

    return sorted(set(filled))


def student_rows(ys):
    if len(ys) < 3:
        return []

    gaps = np.diff(ys)

    valid = [
        g for g in gaps
        if 10 <= g <= 100
    ]

    if not valid:
        return []

    height = np.median(valid)

    rows = []

    for y1, y2 in zip(
        ys[:-1],
        ys[1:],
    ):
        if (
            height * 0.55
            <= y2 - y1
            <= height * 1.6
        ):
            rows.append((y1, y2))

    return rows


def crop(img, x1, y1, x2, y2, pad=2):
    h, w = img.shape[:2]

    x1 = max(0, int(x1 + pad))
    y1 = max(0, int(y1 + pad))

    x2 = min(w, int(x2 - pad))
    y2 = min(h, int(y2 - pad))

    if x2 <= x1 or y2 <= y1:
        return np.empty((0, 0), dtype=np.uint8)

    return img[y1:y2, x1:x2]


# ============================================================
# ATTENDANCE MARK
# ============================================================

def attendance_mark(cell):
    """
    We don't force OCR on tiny handwritten marks.

    If meaningful ink exists -> •
    If empty -> blank

    User can then correct • into U/H/C/etc.
    """

    if cell is None or cell.size == 0:
        return ""

    gray = (
        cv2.cvtColor(cell, cv2.COLOR_RGB2GRAY)
        if cell.ndim == 3
        else cell
    )

    binary = cv2.threshold(
        gray,
        0,
        255,
        cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU,
    )[1]

    h, w = binary.shape

    # Ignore cell borders
    mx = max(2, int(w * 0.18))
    my = max(2, int(h * 0.18))

    inner = binary[
        my:max(my + 1, h - my),
        mx:max(mx + 1, w - mx),
    ]

    if inner.size == 0:
        return ""

    ink = np.mean(inner > 0)

    return "•" if ink > 0.015 else ""


# ============================================================
# EXTRACT ATTENDANCE
# ============================================================

def extract_attendance(img, model, status=None):
    img, gray = preprocess(img)

    xs, ys = detect_grid(img, gray)
    ys = recover_missing_lines(ys)

    days = find_day_columns(xs)

    if days is None or len(days) < 10:
        raise ValueError(
            "Could not detect attendance day columns."
        )

    start = xs.index(days[0])

    if start < 3:
        raise ValueError(
            "Could not locate Name/Sex columns."
        )

    # Columns immediately before day 1
    sex_x = xs[start - 1]
    name_x = xs[start - 2]
    serial_x = xs[start - 3]

    no_x = (
        xs[start - 4]
        if start >= 4
        else xs[0]
    )

    rows = student_rows(ys)

    if not rows:
        raise ValueError(
            "Could not detect student rows."
        )

    # --------------------------------------------------------
    # Crop every text field for every row FIRST, then run OCR
    # once on the whole batch. This replaces calling the full
    # detect+recognize pipeline separately for every single cell
    # (rows * 4 model calls), which is what was making the app sit
    # at "Processing page 1..." for a long time on CPU with no
    # visible progress.
    # --------------------------------------------------------

    if status is not None:
        status.write(f"Detected {len(rows)} student row(s). Cropping fields...")

    crops = []
    field_of = []  # (row_index, field_name), same order as `crops`

    for i, (y1, y2) in enumerate(rows):

        crops.append(crop(img, name_x, y1, sex_x, y2, 3))
        field_of.append((i, "name"))

        crops.append(crop(img, no_x, y1, serial_x, y2, 3))
        field_of.append((i, "no"))

        crops.append(crop(img, serial_x, y1, name_x, y2, 3))
        field_of.append((i, "serial"))

        crops.append(crop(img, sex_x, y1, days[0], y2, 2))
        field_of.append((i, "sex"))

    if status is not None:
        status.write(f"Recognizing text in {len(crops)} cell(s) (single batched pass)...")

    texts = reco_batch(model, crops)

    fields = [{} for _ in rows]

    for (row_i, field_name), text in zip(field_of, texts):
        fields[row_i][field_name] = text

    if status is not None:
        status.write("Building table and reading attendance marks...")

    records = []
    
    # Debug: track which rows are being filtered
    debug_info = []

    for i, (y1, y2) in enumerate(rows):

        name = fields[i].get("name", "")

        # Skip header
        lower = name.lower()

        if (
            "student" in lower
            or "full name" in lower
        ):
            debug_info.append(f"Row {i}: Skipped (header row)")
            continue

        no = fields[i].get("no", "")
        serial = fields[i].get("serial", "")
        sex = fields[i].get("sex", "")

        # Ignore completely empty non-student rows - but be lenient
        # If we have SOME content (name OR serial), include it
        if not name and not serial and not no:
            debug_info.append(f"Row {i}: Skipped (all fields empty)")
            continue

        debug_info.append(f"Row {i}: Included (name='{name}', serial='{serial}', no='{no}')")

        row = {
            "No": no or str(len(records) + 1),
            "Serial No": serial,
            "Student Full Name": name,
            "Sex": sex,
        }

        for j in range(len(days) - 1):

            cell = crop(
                img,
                days[j],
                y1,
                days[j + 1],
                y2,
                1,
            )

            row[str(j + 1)] = attendance_mark(cell)

        records.append(row)

    if not records:
        debug_msg = "\n".join(debug_info) if debug_info else "No rows processed"
        # Fallback: create empty rows for each detected row so user can manually fill in
        if len(rows) > 1:  # At least one real row (header might be included)
            # Create empty records for each detected row, skipping the first one (usually header)
            start_idx = 1 if len(rows) > 2 else 0  # Skip first row if it's likely a header
            for i, (y1, y2) in enumerate(rows[start_idx:], start=start_idx):
                row = {
                    "No": str(i),
                    "Serial No": "",
                    "Student Full Name": "",
                    "Sex": "",
                }
                for j in range(len(days) - 1):
                    cell = crop(
                        img,
                        days[j],
                        y1,
                        days[j + 1],
                        y2,
                        1,
                    )
                    row[str(j + 1)] = attendance_mark(cell)
                records.append(row)
            # Return the records without the student info - user can fill it in manually
            # The warning will be handled by the calling code
        else:
            raise ValueError(
                f"Grid found, but no student rows were extracted.\n\nDEBUG INFO:\n{debug_msg}\n\nPlease check that:\n1. The PDF contains a table with student data\n2. The Name and Serial No columns are present\n3. Student rows are visible in the document"
            )

    df = pd.DataFrame(records)

    # Always create all 31 dates
    for day in range(1, 32):
        if str(day) not in df:
            df[str(day)] = ""

    columns = (
        [
            "No",
            "Serial No",
            "Student Full Name",
            "Sex",
        ]
        + [str(x) for x in range(1, 32)]
    )

    df = df[columns]

    # Debug image
    preview = img.copy()

    for x in days:
        cv2.line(
            preview,
            (x, 0),
            (x, preview.shape[0]),
            (0, 255, 0),
            2,
        )

    for y1, y2 in rows:
        cv2.line(
            preview,
            (0, y1),
            (preview.shape[1], y1),
            (255, 0, 0),
            1,
        )

    return df, preview


# ============================================================
# OLLAMA
# ============================================================

def ollama_cleanup(df, model_name):
    """
    Ollama only examines headers/document structure.
    It does NOT change attendance values.
    """

    preview = df.head(10).to_dict("records")

    prompt = f"""
This is an OCR extracted attendance sheet.

Data:
{json.dumps(preview, ensure_ascii=False)}

Return JSON:

{{
 "document_type":"attendance_sheet",
 "notes":"short description"
}}

Do not rewrite names, dates, IDs or attendance values.
"""

    try:
        r = requests.post(
            OLLAMA_URL,
            json={
                "model": model_name,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt,
                    }
                ],
                "format": "json",
                "stream": False,
                "options": {
                    "temperature": 0
                },
            },
            timeout=1800,
        )

        r.raise_for_status()

        return json.loads(
            r.json()["message"]["content"]
        )

    except Exception as e:
        return {
            "document_type": "attendance_sheet",
            "notes": f"Ollama unavailable: {e}",
        }


# ============================================================
# EXCEL
# ============================================================

def make_excel(df):
    out = io.BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    thin = Side(
        style="thin",
        color="888888",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    fill = PatternFill(
        "solid",
        fgColor="DDEBF7",
    )

    for c, column in enumerate(df.columns, 1):

        cell = ws.cell(
            1,
            c,
            str(column),
        )

        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center"
        )

    for r, row in enumerate(
        df.itertuples(index=False),
        2,
    ):

        for c, value in enumerate(row, 1):

            cell = ws.cell(
                r,
                c,
                str(value),
            )

            cell.border = border
            cell.alignment = Alignment(
                horizontal="center"
            )

    for i, column in enumerate(df.columns, 1):

        if column == "Student Full Name":
            width = 30

        elif str(column).isdigit():
            width = 5

        else:
            width = 12

        ws.column_dimensions[
            get_column_letter(i)
        ].width = width

    ws.freeze_panes = "A2"

    wb.save(out)

    return out.getvalue()


# ============================================================
# PDF
# ============================================================

def make_pdf(
    df,
    title,
    school,
    grade,
    class_name,
    teacher,
    month,
    year,
):
    """
    Recreates attendance as a printable
    A3 landscape PDF with all 31 days.
    """

    out = io.BytesIO()

    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A3),
        leftMargin=5 * mm,
        rightMargin=5 * mm,
        topMargin=5 * mm,
        bottomMargin=5 * mm,
    )

    styles = getSampleStyleSheet()

    story = [
        Paragraph(
            title,
            styles["Title"],
        ),
        Spacer(1, 2 * mm),
    ]

    # --------------------------------------------------------
    # Metadata
    # --------------------------------------------------------

    meta = [
        [
            f"School: {school}",
            f"Grade: {grade}",
            f"Class: {class_name}",
        ],
        [
            f"Teacher: {teacher}",
            f"Month: {month}",
            f"Year: {year}",
        ],
    ]

    mt = Table(
        meta,
        colWidths=[
            125 * mm,
            125 * mm,
            125 * mm,
        ],
    )

    mt.setStyle(
        TableStyle([
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black,
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8,
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE",
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                4,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                4,
            ),
        ])
    )

    story += [
        mt,
        Spacer(1, 3 * mm),
    ]

    # --------------------------------------------------------
    # Ensure columns
    # --------------------------------------------------------

    temp = df.copy()

    for column in [
        "No",
        "Serial No",
        "Student Full Name",
        "Sex",
    ]:
        if column not in temp:
            temp[column] = ""

    for day in range(1, 32):
        if str(day) not in temp:
            temp[str(day)] = ""

    # --------------------------------------------------------
    # Absent totals
    # --------------------------------------------------------

    def absent(row):
        return sum(
            str(row[str(day)]).strip().upper()
            in {"A", "ABSENT"}
            for day in range(1, 32)
        )

    temp["Total Absent Days"] = temp.apply(
        absent,
        axis=1,
    )

    columns = (
        [
            "No",
            "Serial No",
            "Student Full Name",
            "Sex",
        ]
        + [str(x) for x in range(1, 32)]
        + ["Total Absent Days"]
    )

    temp = temp[columns]

    header = (
        [
            "No",
            "Serial No",
            "Student Full Name",
            "Sex",
        ]
        + [str(x) for x in range(1, 32)]
        + ["Total\nAbsent"]
    )

    data = [header]

    for _, row in temp.iterrows():

        data.append([
            "" if pd.isna(row[c])
            else str(row[c])
            for c in columns
        ])

    # Add printable blank rows
    while len(data) < 21:

        blank = [""] * len(columns)

        blank[0] = str(len(data))

        data.append(blank)

    widths = (
        [
            8 * mm,
            22 * mm,
            60 * mm,
            10 * mm,
        ]
        + [7 * mm] * 31
        + [16 * mm]
    )

    table = Table(
        data,
        colWidths=widths,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.lightgrey,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                5.5,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.black,
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER",
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE",
            ),
            (
                "ALIGN",
                (2, 1),
                (2, -1),
                "LEFT",
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                2,
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                2,
            ),
        ])
    )

    story.append(table)

    # --------------------------------------------------------
    # Daily totals
    # --------------------------------------------------------

    totals = [
        "Total Absent Students",
        "",
        "",
        "",
    ]

    for day in range(1, 32):

        count = sum(
            str(x).strip().upper()
            in {"A", "ABSENT"}
            for x in temp[str(day)]
        )

        totals.append(
            str(count) if count else ""
        )

    totals.append("")

    total_table = Table(
        [totals],
        colWidths=widths,
    )

    total_table.setStyle(
        TableStyle([
            (
                "SPAN",
                (0, 0),
                (3, 0),
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.black,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, -1),
                "Helvetica-Bold",
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                5.5,
            ),
            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER",
            ),
        ])
    )

    story += [
        total_table,
        Spacer(1, 3 * mm),
    ]

    # --------------------------------------------------------
    # Legend
    # --------------------------------------------------------

    legend = [
        [
            "Unknown Absence", "U",
            "Illness", "H",
            "Chores", "C",
            "Not Interested", "NI",
        ],
        [
            "No School Materials", "NSM",
            "Other", "O",
            "Holiday", "L",
            "Weekend", "X",
        ],
    ]

    lt = Table(
        legend,
        colWidths=[
            40 * mm,
            12 * mm,
        ] * 4,
    )

    lt.setStyle(
        TableStyle([
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.35,
                colors.black,
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                6,
            ),
            (
                "ALIGN",
                (1, 0),
                (1, -1),
                "CENTER",
            ),
            (
                "ALIGN",
                (3, 0),
                (3, -1),
                "CENTER",
            ),
            (
                "ALIGN",
                (5, 0),
                (5, -1),
                "CENTER",
            ),
            (
                "ALIGN",
                (7, 0),
                (7, -1),
                "CENTER",
            ),
        ])
    )

    story.append(lt)

    doc.build(story)

    return out.getvalue()


# ============================================================
# UI
# ============================================================

with st.sidebar:

    st.header("Settings")

    language_label = st.selectbox(
        "Handwriting / print language",
        list(LANGUAGE_OPTIONS.keys()),
        index=0,
    )

    ocr_lang = LANGUAGE_OPTIONS[language_label]

    st.caption(
        "Sets which PaddleOCR recognition model is loaded — pick the "
        "language the sheet is actually written in, not the UI language."
    )

    model_name = st.text_input(
        "Ollama model",
        OLLAMA_MODEL,
    )

    use_ollama = st.checkbox(
        "Use Ollama",
        True,
    )

    st.caption(
        "Ollama does not modify attendance values."
    )


uploaded = st.file_uploader(
    "Upload attendance image or PDF",
    type=[
        "png",
        "jpg",
        "jpeg",
        "webp",
        "pdf",
    ],
)


if uploaded:

    pages = read_file(uploaded)

    st.success(
        f"{len(pages)} page(s) loaded"
    )

    page = st.selectbox(
        "Preview page",
        range(1, len(pages) + 1),
    )

    st_image_compat(
        pages[page - 1],
    )

    if st_button_compat(
        "🚀 Digitize",
        type="primary",
    ):

        try:

            with st.status(
                "Loading OCR models... (first run downloads pretrained "
                "weights, this can take a couple of minutes)",
                expanded=True,
            ) as status:

                model = get_ocr(ocr_lang)

                status.update(label="OCR models ready.")

                frames = []
                previews = []

                progress = st.progress(0)

                for i, image in enumerate(pages):

                    status.update(
                        label=f"Processing page {i + 1} of {len(pages)}..."
                    )

                    df, preview = extract_attendance(
                        image,
                        model,
                        status=status,
                    )

                    frames.append(df)
                    previews.append(preview)

                    progress.progress(
                        (i + 1) / len(pages)
                    )

                st.session_state["df"] = pd.concat(
                    frames,
                    ignore_index=True,
                )

                st.session_state["preview"] = previews

                # Check if OCR extraction failed to get student info
                df = st.session_state["df"]
                if df is not None and len(df) > 0:
                    empty_names = df["Student Full Name"].fillna("").str.strip().eq("").all()
                    empty_serials = df["Serial No"].fillna("").str.strip().eq("").all()
                    if empty_names and empty_serials:
                        st.warning(
                            "⚠️ **OCR extraction failed for student names and serial numbers.**\n\n"
                            "The attendance grid was detected and attendance marks were extracted, "
                            "but student information could not be recognized. "
                            "Please manually enter the student names and serial numbers below."
                        )

                if use_ollama:

                    status.update(label="Running Ollama cleanup pass...")

                    info = ollama_cleanup(
                        st.session_state["df"],
                        model_name,
                    )

                    st.session_state["ollama"] = info

                status.update(
                    label="Done.",
                    state="complete",
                    expanded=False,
                )

        except Exception as e:

            st.error(f"Extraction failed: {e}")

            # IMPORTANT: gives exact traceback
            st.exception(e)


# ============================================================
# RESULT
# ============================================================

if "df" in st.session_state:

    st.divider()

    if "preview" in st.session_state:

        st.subheader("Detected Grid")

        preview_page = st.selectbox(
            "Grid page",
            range(
                1,
                len(st.session_state["preview"]) + 1,
            ),
        )

        st_image_compat(
            st.session_state["preview"][
                preview_page - 1
            ],
        )

    if "ollama" in st.session_state:

        info = st.session_state["ollama"]

        st.info(
            f"{info.get('document_type', 'attendance_sheet')} — "
            f"{info.get('notes', '')}"
        )

    st.subheader("✏️ Review Attendance")

    edited = st_data_editor_compat(
        st.session_state["df"],
        height=600,
        num_rows="dynamic",
    )

    # ========================================================
    # PDF DETAILS
    # ========================================================

    st.subheader("PDF Information")

    a, b, c = st.columns(3)

    with a:
        school = st.text_input(
            "School Name"
        )

        grade = st.text_input(
            "Grade"
        )

    with b:
        class_name = st.text_input(
            "Class"
        )

        teacher = st.text_input(
            "Teacher"
        )

    with c:
        month = st.text_input(
            "Month"
        )

        year = st.text_input(
            "Year"
        )

    title = st.text_input(
        "PDF Title",
        "ABSENCE RECORDING SHEET",
    )

    # ========================================================
    # EXPORT
    # ========================================================

    excel = make_excel(edited)

    csv = edited.to_csv(
        index=False
    ).encode(
        "utf-8-sig"
    )

    pdf = make_pdf(
        edited,
        title,
        school,
        grade,
        class_name,
        teacher,
        month,
        year,
    )

    x, y, z = st.columns(3)

    with x:

        st_download_button_compat(
            "⬇️ Excel",
            excel,
            "attendance.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with y:

        st_download_button_compat(
            "⬇️ CSV",
            csv,
            "attendance.csv",
            "text/csv",
        )

    with z:

        st_download_button_compat(
            "⬇️ Printable PDF",
            pdf,
            "attendance.pdf",
            "application/pdf",
        )